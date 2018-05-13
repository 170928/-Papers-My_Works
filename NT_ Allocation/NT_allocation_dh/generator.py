import numpy as np
import random
import itertools
import math
import openpyxl
import random
import itertools

class Generator:
    def __init__(self, User, Nt, mPathLoss, SNR):

        self.SNRdB = SNR
        #noise 평상적인 값이 0.0316이 나오도록 한다
        self.noise = (1/10)**(self.SNRdB/10)
        #====================================================
        self.mPathLoss = mPathLoss

        #a+bi가 채널의 형태로 channel coefficient라는 이름으로 존재.
        #실험을 위해 채널을 랜덤으로 만드는 부분

        self.mGd = np.zeros(User*Nt)
        self.Gd = np.zeros(User*Nt)
        self.silr = np.zeros(User*Nt)
        self.real = np.random.normal(0, 1.0, size=[User*Nt])
        self.comp = np.random.normal(0, 1.0, size=[User*Nt])
        self.temp = np.zeros([User*Nt], dtype=complex)

        for i in range(User):
            for j in range(Nt):
                self.temp[i*Nt + j] = (1/(2**(0.5))) * complex( self.real[i*Nt + j], self.comp[i*Nt + j] )

        #mGd = abs(self.temp).^2
        self.mGd = abs(self.temp)
        #================================================================================================
        #g_ij :: 채널의 크기(magnitude)
        #[1x15]형태의 배열로 U1A1.....U3A5 의 논문에서 사용되는 g(i,j)
        self.Gd = self.mGd ** 2
        #================================================================================================


        for q in range(0, User):
            for w in range(0, Nt):
                self.sum = 0
                for e in range(0, Nt):
                    if (q*Nt+w) != (q*Nt+e):
                        self.sum += self.Gd[q*Nt + e]
                self.silr[q * Nt + w] = (self.Gd[q * Nt + w] / ( self.sum ))
        # ================================================================================================
        # 위에서 만든 silr 배열은 [1x15]배열로 U1A1......U3A5 의 논문에서 사용되는 파이(i,j)
        # ================================================================================================



        #======================Input Normalization=====================
        # ================================================================================================
        # norm 배열은 [1x15]배열로 U1A1 ...... U3A5 로 논문에서 Input으로 사용한 값.
        # ================================================================================================
        #print(self.silr)
        #print('===')

        self.norm = np.zeros([User*Nt])

        #self.temp = np.zeros([3])
        #for i in range(0, Nt):
        #    self.temp[0] = self.silr[i]
        #    self.temp[1] = self.silr[Nt + i]
        #    self.temp[2] = self.silr[2*Nt + i]

        #    min = np.min(self.temp)
        #    max = np.max(self.temp)
        #    #print('temp', self.temp[0], ' / ', self.temp[1], ' / ', self.temp[2])
        #    #print('min', min, 'max', max)

         #   self.temp[0] = (self.temp[0] - min)/(max-min)
         #   self.temp[1] = (self.temp[1] - min)/(max-min)
         #   self.temp[2] = (self.temp[2] - min)/(max-min)

            #print('temp   ', self.temp[0], self.temp[1], self.temp[2])

         #   self.norm[i] = self.temp[0]
         #   self.norm[Nt + i] = self.temp[1]
         #   self.norm[2*Nt + i] = self.temp[2]

        #print(self.norm)
        # ================================================================================================
        # self.norm 배열은 [1x15]배열로 U1A1 ...... U3A5 로 논문에서 Input으로 사용한 값.
        # ================================================================================================

        for i in range(0, User):
            for j in range(0, Nt):
                self.norm[i * Nt + j ] = (self.silr[i * Nt + j] - np.mean(self.silr)) / (np.max(self.silr) - np.min(self.silr))



    def factorial(self, n):
        fac = 1
        for i in range(1, n + 1):
            fac *= i
        return fac


    def optimal(self, User, Nt):

        ####### Nt>U?
        #NtCUesr * User!
        # pa.Lfull = factorial(pa.Nt)/factorial(pa.Nt - pa.U);
        # mLabelTmp = perms(1:pa.Nt);
        # mLabelTmp = mLabelTmp(:,1:pa.U);
        # pa.mLabelASCO = unique(mLabelTmp,'rows');

        #self.Lfull = self.factorial(Nt) / self.factorial(Nt - User)
        #self.LabelTemp = np.array(list(itertools.permutations(np.arange(0, Nt, 1))))
        #self.LabelCut = self.LabelTemp[:, 0:User]

        #np.array(list(itertools.permutations(np.arange(0, 5, 1), 3)))

        # ======================================가능한 조합 경우의 수====================================
        self.availableSet = np.array(list(itertools.permutations(np.arange(0,Nt,1),User)))
        # ===============================================================================================

        # ================================================================================================
        # Sumrate가 최대인 해를 구하는 알고리즘===========================================================
        # ================================================================================================
        #Calculate sum_rate of all cases
        self.sumrateSet = np.zeros([self.availableSet.shape[0]])

        for i in range(0,self.availableSet.shape[0]):
            self.tempLabel = self.availableSet[i]
            self.vRu = np.zeros([User])

            for j in range(0,User):
                self.Su = self.Gd[j*Nt + self.tempLabel[j]]
                self.lu = 0
                for k in range(0,User):
                    if k != j:
                        self.lu = self.lu + self.Gd[j*Nt + self.tempLabel[k]]
                self.vRu[j] = math.log2(1 + (self.Su/(self.noise + self.lu)))
                self.sumrateSet[i] = self.sumrateSet[i] + self.vRu[j]


        # ================================================================================================
        # comb 변수에는 Sumrate를 최대로 하는 조합 [ 3 2 1 ]과 같은 애의 index 가 들어있다. 이 index는 self.availableSet[ index ] 를 하면 [ 3 2 1 ] 의 배열을 찾을 수 있게 해준다
        # ================================================================================================
        self.comb = np.argmax(self.sumrateSet)

        # ================================================================================================
        # 최대 sumrate는 self.sumrateSet[ self.comb ] 를 하면 구할 수 있다.
        # ================================================================================================
        #print(self.sumrateSet[self.comb])
        #print(self.sumrateSet)

        #===============학습을 위해 comb를 원핫인코딩한다=====================
        self.label = np.zeros([len(self.availableSet)])
        self.label[self.comb] = 1
        #self.label 은 1x60의 comb 위치만이 1

    def getRandOptVal(self,num_rand=1):
        max = 0
        for i in range(num_rand):
            randpicked = random.randrange(0,len(self.availableSet))
            temp = self.sumrateSet[randpicked]
            if(temp>max):
                max = temp
        return max

    def getGreedyOptVal(self):
        greedyOpt =0
        return greedyOpt

def save():

    User = 3
    Nt = 5

    snr_set = snr_generator()

    label = openpyxl.load_workbook('label.xlsx')
    comb = openpyxl.load_workbook('comb.xlsx')
    silr = openpyxl.load_workbook('silr.xlsx')
    norm = openpyxl.load_workbook('norm.xlsx')

    label_sheet = label.active
    comb_sheet = comb.active
    silr_sheet = silr.active
    norm_sheet = norm.active

    sheet_num = 1
    for s in snr_set:
        # s 는 snr_set에 있는 5, 7 ... 29까지를 하나씩 꺼내온다

        for r in range(0, 10000):

            #genrator 클래스를 인스턴스화 시킬 때 SNR 값을 넣어주면 해당 값에 대응되는 noise를 넣은 Sumrate 최적 값을 찾을 수 있는 환경구성
            a = Generator(3, 5, 1, SNR=s)
            #a.norm 으로 정규화된 input 에 접근
            #a.silr 으로 정규화이전 silr 에 접근
            #a.Gd 로 g(ij) 모음에 접근
            #optimal 함수를 호출하면 최적값을 찾는다.
            #a.comb
            #a.sumrateSet[a.comb]
            #위 의 2 변수를 optimal 다음에 접근하면 Sumrate가 최대인 조합의 index와 그때의 sumrate를 얻을 수 있다.
            a.optimal(3, 5)

            for q in range(0, User*Nt):
                    silr_sheet.cell(row=r + 1, column=q + 1).value = a.silr[q]

            for k in range(0, User):
                label_sheet.cell(row=r + 1, column=k + 1).value = a.availableSet[a.comb, k]

            comb_sheet.cell(row=r + 1, column=1).value = a.comb

            for t in range(0, User*Nt):
                norm_sheet.cell(row=r + 1, column=t + 1).value = a.norm[t]



        # 새로운 Sheet 생성
        print("-----------next snr set--------------")
        label_sheet = label.create_sheet(repr(s + 2), sheet_num)
        comb_sheet = comb.create_sheet(repr(s + 2), sheet_num)
        silr_sheet = silr.create_sheet(repr(s + 2), sheet_num)
        norm_sheet = norm.create_sheet(repr(s + 2), sheet_num)
        sheet_num = sheet_num + 1

    label.save('label.xlsx')
    comb.save('comb.xlsx')
    silr.save('silr.xlsx')
    norm.save('norm.xlsx')
    comb.close()
    label.close()
    silr.close()
    norm.close()


def snr_generator(start = 5, end = 30, interval = 2):
    #end는 미만의 개념이므로 30까지
    #start는 5로하자
    return list(range(start, end, interval))

if __name__ == "__main__":

    print("========Gen========")
    #save()

    a = Generator(3,5,1,15)
    a.optimal(3, 5)
    a.getRandOptVal(1)